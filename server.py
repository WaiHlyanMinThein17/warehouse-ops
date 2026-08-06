"""
server.py — GPD Warehouse Ops Backend
Run: python server.py
Then open: http://localhost:5000
"""

from flask import Flask, request, send_file, jsonify
from flask.helpers import make_response
import io
import os
import tempfile
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import (
    Font, Alignment, Border, Side, PatternFill
)
from openpyxl.utils import get_column_letter
from python_calamine import CalamineWorkbook

app = Flask(__name__, static_folder='.', static_url_path='')


# ── Serve the frontend ────────────────────────────────────────────────────────

@app.route('/')
def index():
    return app.send_static_file('index.html')


# ── Processing Log Generator ──────────────────────────────────────────────────

@app.route('/api/processing-log', methods=['GET'])
def generate_processing_log():
    station = request.args.get('station', 'MCN2')
    month   = int(request.args.get('month', date.today().month))   # 1-indexed
    year    = int(request.args.get('year',  date.today().year))

    month_names = [
        '', 'January', 'February', 'March', 'April', 'May', 'June',
        'July', 'August', 'September', 'October', 'November', 'December'
    ]
    month_name = month_names[month]

    # ── Get all weekdays in the month ─────────────────────────────────────────
    weekdays = []
    d = date(year, month, 1)
    while d.month == month:
        if d.weekday() < 5:  # Mon=0 … Fri=4
            weekdays.append(d)
        d += timedelta(days=1)

    # ── Ordinal suffix ────────────────────────────────────────────────────────
    def ordinal(n):
        if 11 <= (n % 100) <= 13:
            return f'{n}th'
        return f'{n}{["th","st","nd","rd","th"][min(n % 10, 4)]}'

    def fmt_day(d):
        names = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']
        return f'{names[d.weekday()]}, {month_name} {ordinal(d.day)}'

    # ── Build workbook ────────────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = 'Processing Log'

    # Styles
    font_title = Font(name='Aptos Narrow', size=26, bold=True)
    font_hdr   = Font(name='Aptos Narrow', size=18, bold=True)
    font_date  = Font(name='Aptos Narrow', size=18, bold=True)
    align_ctr  = Alignment(horizontal='center', vertical='center')
    thin       = Side(style='thin')
    border     = Border(top=thin, bottom=thin, left=thin, right=thin)
    thick_bot  = Border(
        top=thin, bottom=Side(style='medium'), left=thin, right=thin
    )

    # Column widths (match original)
    col_widths = {
        'A': 5, 'B': 8.57, 'C': 17, 'D': 5,
        'E': 10, 'F': 10, 'G': 10, 'H': 10.29, 'I': 11.14
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # Row 2: Title
    ws.row_dimensions[2].height = 34.5
    ws.merge_cells('B2:H2')
    ws['B2'] = f'>>> {station} {month_name} Processing Log <<<'
    ws['B2'].font      = font_title
    ws['B2'].alignment = align_ctr

    # Row 4: Headers
    ws.row_dimensions[4].height = 24
    ws.merge_cells('A4:C4')
    ws.merge_cells('E4:I4')
    ws['A4'] = 'DATE'
    ws['E4'] = 'Processor'
    for cell in [ws['A4'], ws['E4']]:
        cell.font      = font_hdr
        cell.alignment = align_ctr

    # Row 5: spacer
    ws.row_dimensions[5].height = 24

    # Data rows — grouped by week with blank separator rows
    current_row = 6
    last_week   = None

    for d in weekdays:
        week_num = (d.day - 1) // 7

        # Blank separator row between weeks
        if last_week is not None and week_num != last_week:
            ws.row_dimensions[current_row].height = 24
            current_row += 1

        last_week = week_num

        ws.row_dimensions[current_row].height = 24

        # Merge date cells A:C and processor cells E:I
        ws.merge_cells(f'A{current_row}:C{current_row}')
        ws.merge_cells(f'E{current_row}:I{current_row}')

        date_cell = ws[f'A{current_row}']
        proc_cell = ws[f'E{current_row}']

        date_cell.value     = fmt_day(d)
        date_cell.font      = font_date
        date_cell.alignment = align_ctr

        # Apply borders to all cells in the row (A–I)
        is_friday = d.weekday() == 4
        row_border = thick_bot if is_friday else border

        for col_idx in range(1, 10):  # A=1 to I=9
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = row_border

        # Processor cell style
        proc_cell.font      = font_date
        proc_cell.alignment = align_ctr

        current_row += 1

    # ── Stream to browser ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'GPD_{station}_{month_name}_{year}_ProcessingLog.xlsx'
    response = make_response(send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    ))
    response.headers['Access-Control-Allow-Origin'] = '*'
    return response


# ── B-Row Top Sellers Replenishment ────────────────────────────────────────────
#
# Joins three Business Central exports on Division Item No. to prioritise
# replenishing the best-selling items on the B row (prime bins PB2*):
#   1. movement    — what needs replenishing & how much (raw Movement Worksheet
#                    or the processed "Task List" export)
#   2. items       — the ~59MB / 425k-row catalogue, for Sales (Qty.) ranking
#   3. bincontents — current pick-face qty + min/max per B-row bin
#
# The items file is parsed with python-calamine (~15s) rather than openpyxl
# (~113s); we stream it once and keep only rows whose Division Item No. is a
# candidate from the movement file, to bound memory.

B_ROW_PREFIX = 'PB2'


def _norm_id(v):
    """Canonicalise a Division Item No. so joins line up across exports."""
    if v is None:
        return ''
    if isinstance(v, bool):
        return ''
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return 0.0


def _read_sheet(storage):
    """Save an uploaded file to a temp path and parse its first sheet with
    calamine. Returns (header, rows) where header is a list of trimmed strings."""
    suffix = os.path.splitext(storage.filename or '')[1] or '.xlsx'
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    try:
        storage.save(tmp.name)
        tmp.close()
        wb = CalamineWorkbook.from_path(tmp.name)
        ws = wb.get_sheet_by_name(wb.sheet_names[0])
        data = ws.to_python(skip_empty_area=True)
    finally:
        try:
            os.unlink(tmp.name)
        except OSError:
            pass
    if not data:
        return [], []
    header = [str(h).strip() if h is not None else '' for h in data[0]]
    return header, data[1:]


def _col(header, *names):
    for n in names:
        if n in header:
            return header.index(n)
    return -1


@app.route('/api/brow-topsellers', methods=['POST'])
def brow_topsellers():
    movement    = request.files.get('movement')
    items       = request.files.get('items')
    bincontents = request.files.get('bincontents')

    if not (movement and items and bincontents):
        return jsonify({'error': 'All three files are required: movement, items, bincontents.'}), 400

    top_n_raw = (request.form.get('topN') or '50').strip().lower()
    top_n = None if top_n_raw in ('all', '0', '') else max(1, int(_num(top_n_raw)))

    # 1. Movement worksheet → B-row candidate replenishment lines ────────────────
    hdr, rows = _read_sheet(movement)
    m_div  = _col(hdr, 'Division item No.', 'Division Item No.', 'Division Item No', 'Division item No')
    m_desc = _col(hdr, 'Description')
    m_from = _col(hdr, 'From Bin Code', 'From Bin (Bulk)')
    m_to   = _col(hdr, 'To Bin Code', 'To Bin (Prime)')
    m_sug  = _col(hdr, 'Qty. to Handle', 'Suggested Qty')
    m_avail = _col(hdr, 'Available Qty. to Move', 'Available in Bulk')

    if m_div < 0 or m_to < 0:
        return jsonify({'error': 'Movement file is missing a Division Item No. or prime/To bin column.'}), 400

    # Aggregate by (item, prime bin) — BC often emits several movement lines for
    # the same item/bin; summing keeps one prioritised row per pick face.
    agg = {}
    for r in rows:
        to_bin = str(r[m_to]).strip() if m_to >= 0 and r[m_to] is not None else ''
        if not to_bin.startswith(B_ROW_PREFIX):
            continue
        div = _norm_id(r[m_div])
        if not div:
            continue
        key = (div, to_bin)
        c = agg.get(key)
        if c is None:
            c = agg[key] = {
                'divItemNo':      div,
                'description':    (str(r[m_desc]).strip() if m_desc >= 0 and r[m_desc] is not None else ''),
                'fromBin':        (str(r[m_from]).strip() if m_from >= 0 and r[m_from] is not None else ''),
                'toBin':          to_bin,
                'suggestedQty':   0,
                'availableInBulk': 0,
            }
        c['suggestedQty'] += round(_num(r[m_sug])) if m_sug >= 0 else 0
        # Available-in-bulk is item-level and repeated across lines — take the max.
        c['availableInBulk'] = max(c['availableInBulk'], round(_num(r[m_avail])) if m_avail >= 0 else 0)

    candidates = list(agg.values())
    if not candidates:
        return jsonify({'error': 'No B-row (PB2*) replenishment lines found in the movement file.'}), 400

    candidate_ids = {c['divItemNo'] for c in candidates}

    # 2. Bin Contents → current qty / min / max per B-row (div, bin) ──────────────
    hdr, rows = _read_sheet(bincontents)
    b_bin = _col(hdr, 'Bin Code')
    b_div = _col(hdr, 'Division Item No.', 'Division Item No')
    b_qty = _col(hdr, 'Quantity (Base)')
    b_min = _col(hdr, 'Min. Qty.')
    b_max = _col(hdr, 'Max. Qty.')
    b_desc = _col(hdr, 'ItemDescription', 'Description')

    bin_by_div_bin = {}
    bin_by_div = {}
    for r in rows:
        bin_code = str(r[b_bin]).strip() if b_bin >= 0 and r[b_bin] is not None else ''
        if not bin_code.startswith(B_ROW_PREFIX):
            continue
        div = _norm_id(r[b_div]) if b_div >= 0 else ''
        if div not in candidate_ids:
            continue
        rec = {
            'currentQty': round(_num(r[b_qty])) if b_qty >= 0 else 0,
            'minQty':     round(_num(r[b_min])) if b_min >= 0 else 0,
            'maxQty':     round(_num(r[b_max])) if b_max >= 0 else 0,
            'bin':        bin_code,
            'description': (str(r[b_desc]).strip() if b_desc >= 0 and r[b_desc] is not None else ''),
        }
        bin_by_div_bin[(div, bin_code)] = rec
        bin_by_div.setdefault(div, rec)

    # 3. Items → Sales (Qty.) for candidates only (stream the big file once) ──────
    hdr, rows = _read_sheet(items)
    i_div   = _col(hdr, 'Division Item No.', 'Division Item No')
    i_sales = _col(hdr, 'Sales (Qty.)')
    i_desc  = _col(hdr, 'Description')
    i_qoh   = _col(hdr, 'Quantity on Hand')

    sales_by_div = {}
    if i_div >= 0:
        for r in rows:
            div = _norm_id(r[i_div])
            if div in candidate_ids and div not in sales_by_div:
                sales_by_div[div] = {
                    'salesQty':    round(_num(r[i_sales])) if i_sales >= 0 else 0,
                    'description': (str(r[i_desc]).strip() if i_desc >= 0 and r[i_desc] is not None else ''),
                    'qtyOnHand':   round(_num(r[i_qoh])) if i_qoh >= 0 else 0,
                }
    del rows  # free the big list promptly

    # 4. Merge, classify, rank by sales ──────────────────────────────────────────
    merged = []
    for c in candidates:
        div = c['divItemNo']
        s = sales_by_div.get(div, {})
        b = bin_by_div_bin.get((div, c['toBin'])) or bin_by_div.get(div, {})

        sug = c['suggestedQty']
        avail = c['availableInBulk']
        if sug > 0 and avail < sug:
            status = 'Check Bin'
        elif sug > 0:
            status = 'Replenish'
        else:
            status = 'OK'

        merged.append({
            'divItemNo':       div,
            'description':     c['description'] or s.get('description') or b.get('description') or '',
            'fromBin':         c['fromBin'],
            'toBin':           c['toBin'] or b.get('bin', ''),
            'currentQty':      b.get('currentQty', 0),
            'minQty':          b.get('minQty', 0),
            'maxQty':          b.get('maxQty', 0),
            'suggestedQty':    sug,
            'availableInBulk': avail,
            'qtyOnHand':       s.get('qtyOnHand', 0),
            'salesQty':        s.get('salesQty', 0),
            'status':          status,
        })

    merged.sort(key=lambda x: x['salesQty'], reverse=True)
    for i, row in enumerate(merged):
        row['rank'] = i + 1

    result = merged if top_n is None else merged[:top_n]

    stats = {
        'totalCandidates': len(merged),
        'returned':        len(result),
        'checkBinCount':   sum(1 for r in result if r['status'] == 'Check Bin'),
        'totalSuggested':  sum(r['suggestedQty'] for r in result),
        'matchedSales':    len(sales_by_div),
        'topSeller':       (result[0]['description'] or result[0]['divItemNo']) if result else '',
        'topSellerQty':    result[0]['salesQty'] if result else 0,
    }

    return jsonify({'rows': result, 'stats': stats})


# ── Run ───────────────────────────────────────────────────────────────────────

if __name__ == '__main__':
    print('\n  GPD Warehouse Ops running at http://localhost:5000\n')
    app.run(debug=False, port=5000)